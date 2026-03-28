from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def create_excel(df, person_name):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active

    # ---- MERGED NAME CELL ----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    name_text = f"Name: {person_name}"
    ws.cell(row=1, column=1).value = name_text

    # ---- WRITE DATAFRAME ----
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # ---- AUTO WIDTH ----
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        # Adjust width considering name row
        max_length = max(max_length, len(name_text) // 2)

        ws.column_dimensions[col_letter].width = max_length + 3

    # ---- APPLY ALIGNMENT ----
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(output)
    return output.getvalue()